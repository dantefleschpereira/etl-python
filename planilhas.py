import datetime
from functools import reduce
from pathlib import Path
import re
import string
from typing import Any, List, Tuple, Union

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
import pandas as pd
from unidecode import unidecode
import xlrd


SAFE_DATALAKE_EXPORT_PATHS = (
    '/tmp/',
    '/data/user/projects/',
)


def safe_identifier_name(text: str):
    """Manda um texto maluco e te darei uma string
    que pode ser utilizada como nome de coluna ou arquivo."""
    safe_chars = string.ascii_lowercase + \
        string.ascii_uppercase + string.digits + '_'
    safe_text = unidecode(str(text)).strip()
    safe_text, _ = re.subn('\|t|\n|\s', '_', safe_text)
    safe_text = ''.join(
        [c for c in safe_text if c in safe_chars]).replace('__', '_')
    return safe_text


def cvt_xls_to_xlsx(*args, **kw) -> Workbook:
    """Open and convert XLS file to openpyxl.workbook.Workbook object

    from https://stackoverflow.com/questions/9918646/how-to-convert-xls-to-xlsx/42574983#42574983

    @param args: args for xlrd.open_workbook
    @param kw: kwargs for xlrd.open_workbook
    @return: openpyxl.workbook.Workbook


    You need -> from openpyxl.utils.cell import get_column_letter
    """

    book_xls = xlrd.open_workbook(
        *args, formatting_info=True, ragged_rows=True, **kw)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index in range(len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])

        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

        for crange in sheet_xls.merged_cells:
            rlo, rhi, clo, chi = crange

            sheet_xlsx.merge_cells(
                start_row=rlo + 1, end_row=rhi,
                start_column=clo + 1, end_column=chi,
            )

        def _get_xlrd_cell_value(cell):
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                value = datetime.datetime(*xlrd.xldate_as_tuple(value, 0))

            return value

        for row in range(sheet_xls.nrows):
            sheet_xlsx.append((
                _get_xlrd_cell_value(cell)
                for cell in sheet_xls.row_slice(row, end_colx=sheet_xls.row_len(row))
            ))

        for rowx in range(sheet_xls.nrows):
            if sheet_xls.rowinfo_map[rowx].hidden != 0:
                sheet_xlsx.row_dimensions[rowx+1].hidden = True
        for coly in range(sheet_xls.ncols):
            if sheet_xls.colinfo_map[coly].hidden != 0:
                coly_letter = get_column_letter(coly+1)
                sheet_xlsx.column_dimensions[coly_letter].hidden = True
    return book_xlsx


def export_csv_datalake(df: pd.DataFrame, path: Union[str, Path]):
    """Salva um Pandas DataFrame uniformemente, ou seja:
    - UTF8
    - Quotechar '"'
    - Separador ','
    - Quebra de linha '\\n'

    Args:
        df (pd.DataFrame): DataFrame a salvar.
        path (Union[str, Path]): Onde salvar o arquivo.

    Raises:
        Exception: Lança exceção quando se tenta salvar fora dos diretórios considerados "SAFE".
    """
    path_normalized = str(path).lower()
    is_safe_path = reduce(lambda prev_result, curr_path: prev_result or path_normalized.startswith(
        curr_path), SAFE_DATALAKE_EXPORT_PATHS, False)
    if not is_safe_path:
        raise Exception(
            f'{path_normalized} está fora dos diretórios permitidos para uso: {SAFE_DATALAKE_EXPORT_PATHS}.')
    Path(path_normalized).parent.absolute().mkdir(exist_ok=True, parents=True)
    new_columns = {curr_column: safe_identifier_name(
        curr_column) for curr_column in df.columns}
    df = df.rename(columns=new_columns)
    df.to_csv(path_normalized, index=False, encoding='utf8',
              sep=',', quotechar='"', line_terminator='\n')
    print(f'Salvo {path_normalized}')


def export_parquet_datalake(df: pd.DataFrame, path: Union[str, Path]):
    """Salva um Pandas DataFrame uniformemente em formato Parquet.

    Args:
        df (pd.DataFrame): DataFrame a salvar.
        path (Union[str, Path]): Onde salvar o arquivo.

    Raises:
        Exception: Lança exceção quando se tenta salvar fora dos diretórios considerados "SAFE".
    """
    path_normalized = str(path).lower()
    is_safe_path = reduce(lambda prev_result, curr_path: prev_result or path_normalized.startswith(
        curr_path), SAFE_DATALAKE_EXPORT_PATHS, False)
    if not is_safe_path:
        raise Exception(
            f'{path_normalized} está fora dos diretórios permitidos para uso: {SAFE_DATALAKE_EXPORT_PATHS}.')
    Path(path_normalized).parent.absolute().mkdir(exist_ok=True, parents=True)
    new_columns = {curr_column: safe_identifier_name(
        curr_column) for curr_column in df.columns}
    df = df.rename(columns=new_columns)
    try:
        df.to_parquet(path_normalized, index=False, version='1.0')
    except ValueError:
        df.pop(df.columns[-1])
        df.to_parquet(path_normalized, index=False, version='1.0')
    print(f'Salvo {path_normalized}')


def get_xlsx_column_headers(
        xslx_path: Union[str, Path], sheet_name: Union[str, int] = None,
        min_row=None, max_row=None, rows_ffill=0, min_col=None, strftime='%d/%m/%Y, %H:%M:%S',
) -> Tuple[List[str], str]:
    """Normaliza nomes de colunas de uma planilha XLSX para ser utilizado em DataFrame pandas.

    Args:
        xslx_path (Union[str, Path]): Arquivo Excel.
        sheet_name (Union[str, int], optional): Nome da planilha (aba). Defaults to None.
        min_row (_type_, optional): Linha onde inicia o cabeçalho. Defaults to None.
        max_row (_type_, optional): Linha onde termina o cabeçalho. Defaults to None.
        rows_ffill (int, optional): Quantidade de linhas a partir da primeira em que se deve repetir o valor caso a próxima coluna seja vazia. Defaults to 0.
        strftime (str, optional): STRFTIME string to format if cell is datetime. Defaults to '%d/%m/%Y, %H:%M:%S'.

    Returns:
        Tuple[List[str], str]: _description_
    """
    retval = []
    worksheet_name = sheet_name or 0
    with open(xslx_path, 'rb') as xlsx_file:
        wb = load_workbook(xlsx_file, read_only=True,
                           data_only=True, keep_links=False,
                           )
        try:
            ws = wb[worksheet_name]
        except KeyError:
            ws = wb.worksheets[worksheet_name]

        worksheet_name = sheet_name or ws.title

        linhas = []
        row_count = min_row
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, values_only=True):
            colunas = []
            prev_cell_val = None
            for cell in row:
                should_ffill = row_count - min_row < rows_ffill
                try:
                    cell = cell.strftime(strftime)
                except:
                    cell = cell or ''
                final_cell_val = prev_cell_val if not cell and should_ffill else str(
                    cell)
                prev_cell_val = final_cell_val
                colunas.append(final_cell_val or '')
            linhas.append(colunas)
            row_count += 1
        colunas_agrupadas = zip(*linhas)
        for grupo_coluna in colunas_agrupadas:
            coluna_str = '_'.join(filter(None, grupo_coluna))
            coluna_safe = safe_identifier_name(coluna_str)
            retval.append(coluna_safe)
    return retval, worksheet_name


def get_xlsx_values(xslx_path: Union[str, Path], sheet_name: Union[str, int] = None, min_row=None, max_col=None, min_col=None) -> List[List[Any]]:
    """Extrai uma matriz que corresponde aos dados de uma planilha.

    Args:
        xslx_path (Union[str, Path]): Arquivo Excel.
        sheet_name (Union[str, int], optional): Nome da planilha (aba). Defaults to None.
        min_row (_type_, optional): Linha onde os dados começam. Defaults to None.
        max_col (_type_, optional): Linha onde os dados terminam. Defaults to None.

    Returns:
        List[List[Any]]: Matrix com dados lidos.
    """
    retval = []
    worksheet_name = sheet_name or 0
    with open(xslx_path, 'rb') as xlsx_file:
        wb = load_workbook(xlsx_file, read_only=True,
                           data_only=True, keep_links=False,
                           )
        try:
            ws = wb[worksheet_name]
        except KeyError:
            ws = wb.worksheets[worksheet_name]

        retval = []
        empty_rows = 0
        for row in ws.iter_rows(min_row=min_row, max_col=max_col, min_col=min_col, values_only=True):
            colunas = []
            empty_cols = 0
            for cell in row:
                if not cell:
                    empty_cols += 1
                    cell = ''
                colunas.append(cell)
            empty_rows = empty_rows + 1 if empty_cols == len(colunas) else 0
            if empty_rows > 2:
                break
            retval.append(colunas)
    return retval


def transform_xlsx_dataframe(
        xlsx_file: Path, data_startrow=9, data_startcolumn=None,
        header_startrow=6, header_nrows=3, header_rows_ffill=0,
        sheet_name: str = None, strftime='%d/%m/%Y, %H:%M:%S',) -> Tuple[pd.DataFrame, str]:
    """Transforma uma planiha dentro de um "workbook" Excel em um DataFrame pandas com colunas normalizadas.

    Args:
        xlsx_file (Path): Caminho do arquivo Excel.
        data_startrow (int, optional): Linha em que começam os dados. Defaults to 9.
        header_startrow (int, optional): Linha em que começa o cabeçalho. Defaults to 6.
        header_nrows (int, optional): Tamanho do cabeçalho em linhas. Defaults to 3.
        header_rows_ffill (int, optional): Quantidade de linhas do cabeçalho que devem ser replicadas à direita caso a próxima coluna seja vazia. Defaults to 0.
        sheet_name (str, optional): Nome ou índice numérico da planilha (aba) a ler. Defaults to None.
        strftime (str, optional): STRFTIME string to format if cell is datetime. Defaults to '%d/%m/%Y, %H:%M:%S'.

    Returns:
        Tuple[pd.DataFrame, str]: Retorna o dataframe e o nome da planilha (caso se tenha informado o índice numérico dela).
    """
    print(f'Processando {xlsx_file}.')
    colunas, final_sheet_name = get_xlsx_column_headers(
        xlsx_file, sheet_name=sheet_name,
        min_col=data_startcolumn,
        min_row=header_startrow, max_row=header_startrow+header_nrows,
        rows_ffill=header_rows_ffill, strftime=strftime,)

    linhas = get_xlsx_values(
        xlsx_file, min_col=data_startcolumn, min_row=data_startrow, sheet_name=sheet_name)

    retval = pd.DataFrame(linhas, dtype=str, columns=colunas)
    print(f'Processado {xlsx_file}.')
    return retval, final_sheet_name


def export_csv_default(csv_files: List[Path], path: Union[str, Path], sep=';', encoding='cp1252'):
    """Unifica uma lista de CSVs em único CSV padronizado.

    Args:
        csv_files (List[Path]): Lista de arquivos.
        path (Union[str, Path]): Nome do arquivo de destino.
        sep (str, optional): Separador utilizado nos CSV lidos. Defaults to ';'.
        encoding (str, optional): Encoding utilizado nos CSV lidos. Defaults to 'cp1252'.
    """
    all_csv = map(lambda x: pd.read_csv(
        x, sep=sep, encoding=encoding, dtype=str), csv_files)
    df = pd.concat(all_csv, ignore_index=True)
    export_csv_datalake(df, path)


def export_parquet_default(csv_files: List[Path], path: Union[str, Path], sep=';', encoding='cp1252'):
    """Unifica uma lista de CSVs em único Parquet padronizado.

    Args:
        csv_files (List[Path]): Lista de arquivos.
        path (Union[str, Path]): Nome do arquivo de destino.
        sep (str, optional): Separador utilizado nos CSV lidos. Defaults to ';'.
        encoding (str, optional): Encoding utilizado nos CSV lidos. Defaults to 'cp1252'.
    """
    all_csv = map(lambda x: pd.read_csv(
        x, sep=sep, encoding=encoding, dtype=str), csv_files)
    df = pd.concat(all_csv, ignore_index=True)
    export_parquet_datalake(df, path)


if __name__ == '__main__':
    print(safe_identifier_name('município'))
